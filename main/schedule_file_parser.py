from collections import defaultdict
from datetime import datetime, time, timedelta
import re
import string
import os
from typing import List
from difflib import SequenceMatcher

from openpyxl import load_workbook
from chunkator import chunkator

from main import models
from moodle.models import MdlCourse

ALPHABET_LIST = list(string.ascii_uppercase)

FACULTY_NAME_CELL = 'Q1'

TITLE_COLUMNS = 2 # titles, weekday names, lessons number fit here
TITLE_ROWS = 6 # faculty, course, group

COURSES_ROW = 3

SPECIALTY_ROW = 4

GROUP_ROW = 5

DAYS_START_ROW = 7

LESSON_NUMBER_COLUMN = 2

MAX_COL = None # horisontal limit
MAX_ROW = None # vertical limit

# LESSON_TIME = {1: (time(hour=8, minute=30), time(hour=9, minute=50)),
#                2: (time(hour=10, minute=10), time(hour=11, minute=30)),
#                3: (time(hour=11, minute=50), time(hour=13, minute=10)),
#                4: (time(hour=14, minute=0), time(hour=15, minute=20)),
#                5: (time(hour=15, minute=40), time(hour=17, minute=0)),
#                6: (time(hour=17, minute=20), time(hour=18, minute=40)),
#                7: (time(hour=19, minute=0), time(hour=20, minute=20)),}

REDUCED_GROUP_STRINGS = ('с.т.', "ст", "ск")

LESSON_NAME_SIMILARITY_THRESHOLD = 0.5

days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

def remove_brackets(string):
    pattern = r"\([^()]*\)"
    result = re.sub(pattern, "", string)
    return result

def find_course(name):
    course = None
    course = MdlCourse.objects.filter(shortname__contains=name).first()
    if not course:
        max_similarity = 0
        for it_course in chunkator(MdlCourse.objects.all(), 100):
            # search course by partial match
            shortname = remove_brackets(it_course.shortname)
            fullname = remove_brackets(it_course.fullname)
            shortname_similarity = SequenceMatcher(None, shortname, name).ratio()
            fullname_similarity = SequenceMatcher(None, fullname, name).ratio()
            similarity = max(shortname_similarity, fullname_similarity)
            if similarity < LESSON_NAME_SIMILARITY_THRESHOLD or \
               similarity < max_similarity:
                continue
            max_similarity = similarity
            course = it_course
    return course

class Group:

    def __init__(self, course_name, spec_name, number_info, faculty):
        self.course_name = course_name
        self.specialty = self.get_specialty(spec_name, faculty)
        self.parse_number_info(number_info) # define type and number
        self.year = self.get_course_year(course_name)
        self.faculty = faculty
        self.db_object = self.get_db_object()

    def __hash__(self) -> int:
        return hash(self.__str__())

    def get_db_object(self):
        try:
            return models.Group.objects.get(year=self.year, specialty=self.specialty,
                                            number=self.number)
        except models.Group.DoesNotExist:
            db_object = models.Group(year=self.year, type=self.type,
                                     specialty=self.specialty, number=self.number)
            db_object.save()
            return db_object

    def parse_number_info(self, number_info):
        self.type = models.Group.Type.BACHELOR
        if any([part in number_info for part in REDUCED_GROUP_STRINGS]):
            self.type = models.Group.Type.REDUCED

        number = ''
        for symbol in number_info:
            if symbol.isdigit():
                number += symbol
            elif number: #number already exist, but symbol is not digit
                break

        self.number = int(number)

        dot_index = number_info.find('.')
        if dot_index > 0 and len(number_info) > dot_index:
            try:
                self.number = int(number_info[dot_index + 1])
            except ValueError: # cannot convert to int
                pass

        open_bracket_index = number_info.find('(')
        if open_bracket_index > 0 and len(number_info) > open_bracket_index:
            try:
                self.number = int(number_info[open_bracket_index + 1])
            except ValueError: # cannot convert to int
                pass


    @staticmethod
    def get_specialty(spec_name, faculty):
        try:
            return models.Specialty.objects.get(faculty=faculty, code=spec_name)
        except models.Specialty.DoesNotExist:
            spec = models.Specialty(faculty=faculty, code=spec_name)
            spec.save()
            return spec

    @staticmethod
    def get_course_year(course_name):
        m = re.search(r"\d", course_name)
        if m:
            year = datetime.now().year - int(m.group(0))
            if datetime.now().month > 10: # second semester
                year += 1
            return year

    def __repr__(self) -> str:
        return self.__str__()

    def __str__(self) -> str:
        return self.db_object.name

class Lesson:

    def __init__(self, week_day: str, number: str, info: str, freq, semester):
        self.week_day = week_day
        self.lesson_number = models.LessonNumber.objects.get(lesson_number=number)
        self.subject, self.location = self.parse_info(info)
        self.freq = freq
        self.semester = semester
        self.groups = []

    @staticmethod
    def parse_info(info):
        m = re.match(r'^((.|\n)+)\s+(\d+(\s+)?к\..+)$', info, re.M)

        if not m:
            raise ValueError(f'Cannot parse lesson info: {info}')

        name, location = m.group(1), m.group(3)
        name = name.replace('\n', ' ')
        try:
            subject = models.Subject.objects.get(title=name)
            if not subject.course_id:
                course = find_course(name)
                if course:
                    subject.course_id = course.id
                    subject.save()
        except models.Subject.DoesNotExist:
            course = find_course(name)
            new_subject = models.Subject(title=name)
            if course:
                new_subject.course_id=course.id
            # TODO: add course url
            new_subject.save()
            subject = new_subject

        return subject, location

    def serialize_to_db(self):
        print(f'serializing lesson {self.__str__()}')
    #     return
        if not self.semester:
            print("Error!! semester not found")
            return
        if not self.groups:
            print("Error!! groups not found")
            return

        db_lesson = (models.Lesson.objects.
                     filter(subject=self.subject, dayofweek=self.week_day,
                            lesson_number=self.lesson_number, semester=self.semester)
                     .first())
        if not db_lesson:
            db_lesson = models.Lesson() # create new
        db_lesson.subject = self.subject
        # db_lesson.starttime, db_lesson.endtime = LESSON_TIME[self.number]
        db_lesson.dayofweek = self.week_day
        db_lesson.weekfrequency = self.freq
        db_lesson.lesson_number = self.lesson_number
        # db_lesson.startdate = startdate
        # db_lesson.enddate = self.semester.enddate
        db_lesson.semester = self.semester
        db_lesson.location = self.location
        db_lesson.save()

        for group in self.groups:
            db_lesson.groups.add(group.db_object)

    # def __eq__(self, __o: object) -> bool:
    #     return (self.info == __o.info and self.week_day == __o.week_day and
    #             self.number == __o.number)

    def __hash__(self) -> int:
        return hash(self.__str__())

    def __repr__(self) -> str:
        return self.__str__()

    def __str__(self):
        name = self.subject.title
        if self.freq == models.WeekFrequency.NUMERATOR:
            name += ' (чисельник)'
        if self.freq == models.WeekFrequency.DENOMINATOR:
            name += ' (знаменник)'
        return f'weekday: {self.week_day} lesson: {self.lesson_number} title: {self.subject.title}'

class ScheduleFileParser:

    def __init__(self, path, faculty, semester) -> None:
        wb = load_workbook(path)
        self.ws = wb.active
        self.faculty = faculty
        self.semester = semester

        self.define_week_days_ranges()
        self.lessons = self.parse_lessons()

    def define_week_days_ranges(self):
        self.week_day_ranges = dict()
        week_day = 0
        for cols in self.ws.iter_cols(min_row=DAYS_START_ROW, min_col=1,
                                      max_row=None, max_col=1):
            for cell in cols:
                if not cell.value:
                    continue
                merged_range = self.get_merged_range(cell.coordinate)
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    self.week_day_ranges[row] = week_day
                week_day += 1

    def parse_group(self, column: int) -> Group | None:
        # print(column)
        course_cell = self.ws.cell(COURSES_ROW, column)
        course_name = self.get_cell_value(course_cell)
        if not course_name:
            return None

        speciality_cell = self.ws.cell(SPECIALTY_ROW, column)
        speciality_name = self.get_cell_value(speciality_cell)
        if not speciality_name:
            return None

        group_cell = self.ws.cell(GROUP_ROW, column)
        group_number = self.get_cell_value(group_cell)
        if not group_number:
            return None

        return Group(course_name, speciality_name, group_number, self.faculty)


    def parse_lessons(self) -> List[Lesson]:
        lessons = list()
        for column_cells in self.ws.iter_cols(min_row=TITLE_ROWS + 1, min_col=TITLE_COLUMNS + 1,
                                      max_row=MAX_ROW, max_col=MAX_COL):
            if not column_cells:
                continue

            group = self.parse_group(column_cells[0].column)
            if not group:
                print(f'Error!! group not found for column {column_cells[0].column}')
                continue

            for cell in column_cells:
                lesson_info = self.get_cell_value(cell)
                if not lesson_info:
                    continue

                cell_merged_range = self.get_merged_range(cell.coordinate)
                if (not cell.value and cell_merged_range and
                    cell_merged_range.min_row != cell.row):
                    continue # avoid lesson duplicates (vertical merged cells)

                week_day = self.week_day_ranges[cell.row]
                if week_day is None:
                    print(f'Error!! week day not found for lesson in cell {cell.row}')
                    continue

                lesson_number_cell = self.ws.cell(cell.row, LESSON_NUMBER_COLUMN)
                if not lesson_number_cell:
                    print(f'Error!! lesson number cell not found for lesson in cell {cell.coordinate}')
                    continue

                # check NUMERATOR or DENOMINATOR
                freq = models.WeekFrequency.EACH_WEEK # from frequency
                if (not cell_merged_range or
                    cell_merged_range.min_row == cell_merged_range.max_row):
                    # if cell occupies only one row it meens lesson is not in each week
                    freq = models.WeekFrequency.DENOMINATOR if lesson_number_cell.value is None \
                        else models.WeekFrequency.NUMERATOR

                lesson_number = self.get_cell_value(lesson_number_cell)
                if lesson_number is None:
                    print(f'Error!! lesson number not found for lesson in cell {cell.coordinate}')
                    continue

                try:
                    lesson = Lesson(week_day, lesson_number, lesson_info, freq, self.semester)
                except ValueError as e:
                    print(f'Warning!! {e}')
                    continue

                try:
                    existed_lesson = next(x for x in lessons if x == lesson)
                    existed_lesson.groups.append(group)
                except StopIteration:
                    lesson.groups.append(group)
                    lessons.append(lesson)

        return lessons

    def serialize_to_db(self):
        for lesson in self.lessons:
            lesson.serialize_to_db()

    def get_cell_value(self, cell):
        if not cell:
            return None
        value = cell.value
        if value:
            return value
        merged_range = self.get_merged_range(cell.coordinate)
        if merged_range:
            value = merged_range.title
            if value is None:
                title_cell = self.ws.cell(merged_range.min_row, merged_range.min_col)
                return title_cell.value
            return value
        return None

    def get_cell_column_range(self, cell):
        if not cell:
            return None
        merged_range = self.get_merged_range(cell.coordinate)
        if merged_range:
            return merged_range.min_col, merged_range.max_col
        else:
            return cell.column, cell.column

    def get_merged_range(self, cell_name):
        if not cell_name:
            return None
        # print(cell_name)
        for merged_range in self.ws.merged_cells.ranges:
            if cell_name in merged_range:
                return merged_range
        return None

    def is_cell_merged(self, cell_name):
        range = self.get_merged_range(cell_name)
        return range is not None
